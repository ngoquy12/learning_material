import openpyxl
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

def export_quiz_to_excel(questions: List[Dict[str, Any]], file_path: str):
    """
    Exports a list of quiz questions to a standardized E-learning Excel format.
    Ensures 13 columns: STT, question_content, answer_1, explanation_answer_1, ...
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Workbook has no active worksheet.")
    ws.title = "Quiz Bank"
    
    # Columns schema
    headers = [
        "STT", "question_content", 
        "answer_1", "explanation_answer_1",
        "answer_2", "explanation_answer_2",
        "answer_3", "explanation_answer_3",
        "answer_4", "explanation_answer_4",
        "isCorrect", "difficulty", "category"
    ]
    
    # Styles
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Write rows
    for row_idx, q in enumerate(questions, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align # STT
        ws.cell(row=row_idx, column=2, value=q.get("question_content", "")).alignment = left_align
        ws.cell(row=row_idx, column=3, value=q.get("answer_1", "")).alignment = left_align
        ws.cell(row=row_idx, column=4, value=q.get("explanation_answer_1", "")).alignment = left_align
        ws.cell(row=row_idx, column=5, value=q.get("answer_2", "")).alignment = left_align
        ws.cell(row=row_idx, column=6, value=q.get("explanation_answer_2", "")).alignment = left_align
        ws.cell(row=row_idx, column=7, value=q.get("answer_3", "")).alignment = left_align
        ws.cell(row=row_idx, column=8, value=q.get("explanation_answer_3", "")).alignment = left_align
        ws.cell(row=row_idx, column=9, value=q.get("answer_4", "")).alignment = left_align
        ws.cell(row=row_idx, column=10, value=q.get("explanation_answer_4", "")).alignment = left_align
        ws.cell(row=row_idx, column=11, value=q.get("isCorrect", 1)).alignment = center_align
        ws.cell(row=row_idx, column=12, value=q.get("difficulty", 1)).alignment = center_align
        ws.cell(row=row_idx, column=13, value=q.get("category", "BÀI MỚI")).alignment = center_align
        
        for col_idx in range(1, 14):
            ws.cell(row=row_idx, column=col_idx).border = thin_border
            
    # Auto-adjust column widths with safety limits
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            # Simple length heuristic
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column or 1)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)
        
    abs_path = os.path.abspath(file_path)
    if os.name == 'nt' and not abs_path.startswith('\\\\?\\'):
        abs_path = '\\\\?\\' + abs_path
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    wb.save(abs_path)

def export_lesson_quiz_to_excel(questions: List[Dict[str, Any]], file_path: str):
    """
    Exports lesson quiz questions to Excel strictly following requirements/Format_Quizz_Lesson.xlsx format:
    4 columns: questionName | point | answerName | isCorrect
    Each question spans N option rows. Only the first option row contains questionName and point.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Workbook has no active worksheet.")
    ws.title = "Quizz Questions"
    
    headers = ["questionName", "point", "answerName", "isCorrect"]
    
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    row_idx = 2
    for q in questions:
        question_text = q.get("question") or q.get("question_content") or ""
        point_val = q.get("point", 20)
        options = q.get("options")
        if not options or not isinstance(options, list):
            options = [
                q.get("answer_1", ""),
                q.get("answer_2", ""),
                q.get("answer_3", ""),
                q.get("answer_4", "")
            ]
        correct_idx = q.get("correct_option_index", 0)
        
        for opt_idx, opt_text in enumerate(options):
            if opt_idx == 0:
                ws.cell(row=row_idx, column=1, value=question_text).alignment = left_align
                ws.cell(row=row_idx, column=2, value=point_val).alignment = center_align
            else:
                ws.cell(row=row_idx, column=1, value=None)
                ws.cell(row=row_idx, column=2, value=None)
                
            ws.cell(row=row_idx, column=3, value=str(opt_text)).alignment = left_align
            ws.cell(row=row_idx, column=4, value="TRUE" if opt_idx == correct_idx else "").alignment = center_align
            
            for col_i in range(1, 5):
                ws.cell(row=row_idx, column=col_i).border = thin_border
            row_idx += 1
            
    ws.column_dimensions['A'].width = 65
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    
    abs_path = os.path.abspath(file_path)
    if os.name == 'nt' and not abs_path.startswith('\\\\?\\'):
        abs_path = '\\\\?\\' + abs_path
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    wb.save(abs_path)
