import sys
import io
import os
import json
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

# Load environment variables from .env
load_dotenv()

from core.state import AgentState
from core.graph import compile_learning_content_workflow

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', write_through=True)

import re
def sanitize_folder_name(name: str) -> str:
    """Loại bỏ các ký tự không hợp lệ cho tên thư mục trên mọi hệ điều hành và an toàn cho URL."""
    name = name.replace("&", "va").replace("%", "").replace("^", "").replace("#", "")
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()

def get_or_rename_sanitized_folder(parent_dir: Path, prefix: str, full_name: str) -> Path:
    """
    Looks for an existing folder starting with prefix (e.g. 'Session 01').
    If found and it has a different name than the new sanitized full_name, renames it.
    This prevents duplicate session/lesson folders when titles change.
    """
    target_name = sanitize_folder_name(full_name)
    target_path = parent_dir / target_name
    
    if parent_dir.exists():
        for item in parent_dir.iterdir():
            if item.is_dir():
                if item.name == target_name:
                    return target_path
                # Prevent matching "Session 010" when prefix is "Session 01"
                if item.name == prefix or item.name.startswith(prefix + " ") or item.name.startswith(prefix + "-"):
                    print(f"  [Auto-Rename] Đồng bộ thư mục cũ: '{item.name}' -> '{target_name}'")
                    try:
                        item.rename(target_path)
                    except Exception as e:
                        print(f"  [Warning] Không thể đổi tên thư mục cũ {item.name}: {e}")
                    return target_path
    return target_path

def parse_all_sessions(excel_path: str):
    """
    Parses all sessions and their corresponding lessons
    from the PM software engineering spreadsheet using dynamic column mapping.
    """
    wb = openpyxl.load_workbook(excel_path)
    
    target_sheet = "Chương trình đào tạo chi tiết"
    ws = wb[target_sheet] if target_sheet in wb.sheetnames else wb.active
    
    # 1. Identify header mapping dynamically to prevent breakage if columns move
    headers = {}
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=0):
        if cell.value:
            val = str(cell.value).strip().lower()
            if "session" in val and "nội dung" not in val: headers["session"] = col_idx
            elif "nội dung session" in val or "chủ đề" in val: headers["session_content"] = col_idx
            elif "chi tiết" in val or "nội dung bài" in val: headers["details"] = col_idx
            elif "lesson" in val or "bài học" in val: headers["lesson"] = col_idx
            elif "sản phẩm" in val or "output" in val: headers["output"] = col_idx
            
    # Fallback to hardcoded indexes if headers weren't found clearly
    h_sess = headers.get("session", 2)
    h_s_content = headers.get("session_content", 3)
    h_lesson = headers.get("lesson", 4)
    h_details = headers.get("details", 5)
    h_out = headers.get("output", 6)
    
    sessions = []
    current_session = None
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue # Skip empty rows
        
        session_val = str(row[h_sess]).strip() if h_sess < len(row) and row[h_sess] else ""
        content_val = str(row[h_s_content]).strip() if h_s_content < len(row) and row[h_s_content] else ""
        lesson_val = str(row[h_lesson]).strip() if h_lesson < len(row) and row[h_lesson] else ""
        details_val = str(row[h_details]).strip() if h_details < len(row) and row[h_details] else ""
        output_val = str(row[h_out]).strip() if h_out < len(row) and row[h_out] else ""
        
        # If a new Session is found
        if session_val.startswith("Session"):
            # Check if we already have it
            existing = [s for s in sessions if s["session_id"] == session_val]
            if existing:
                current_session = existing[0]
            else:
                current_session = {
                    "session_id": session_val,
                    "title": content_val,
                    "lessons": []
                }
                sessions.append(current_session)
            
        # Add lessons under the current session
        if lesson_val.startswith("Lesson") and current_session:
            # Try to safely split Lesson ID and Title
            if ":" in lesson_val:
                l_id, l_title = lesson_val.split(":", 1)
            elif "-" in lesson_val:
                l_id, l_title = lesson_val.split("-", 1)
            else:
                l_id, l_title = lesson_val, lesson_val
                
            current_session["lessons"].append({
                "lesson_id": l_id.strip(),
                "title": l_title.strip(),
                "details": details_val,
                "expected_output": output_val
            })
            
    return sessions

def get_context_curriculum(sessions: list, current_session_id: str) -> list:
    """
    To prevent LLM context window overflow, this function returns a truncated version
    of the curriculum containing only the previous, current, and next session.
    """
    idx = next((i for i, s in enumerate(sessions) if s["session_id"] == current_session_id), -1)
    if idx == -1: return sessions
    
    start = max(0, idx - 1)
    end = min(len(sessions), idx + 2)
    return sessions[start:end]

def initialize_skeleton_structure(sessions, course_dir: Path, requested_parts: list, requested_session: str):
    print("\n=====================================================================")
    print(">>> GIAI ĐOẠN 1.1: KHỞI TẠO KHUNG CẤU TRÚC THƯ MỤC & FILE RỖNG <<<")
    print("=====================================================================")
    for session in sessions:
        session_id = session["session_id"]
        session_title = session.get("title", "")
        req_sess_list = [r.strip() for r in requested_session.split(",")] if requested_session != "all" else ["all"]
        if "all" not in req_sess_list and not any(rs in session_id.lower() for rs in req_sess_list):
            continue

        session_dir = get_or_rename_sanitized_folder(course_dir, session_id, f"{session_id} - {session_title}")
        session_dir.mkdir(parents=True, exist_ok=True)
        print(f"  [Folder] Khởi tạo session: {session_id} - {session_title}")

        is_project_or_hackathon = any(kw in session_title.lower() for kw in ["hackathon", "project", "đồ án", "dự án", "mini project"])
        is_practice = "thực hành" in session_title.lower()

        if is_project_or_hackathon:
            # Ràng buộc cho Mini project: chỉ có 3 thư mục này
            (session_dir / "Bài kiểm tra đầu giờ").mkdir(parents=True, exist_ok=True)
            (session_dir / "Tài liệu đặc tả SRS").mkdir(parents=True, exist_ok=True)
            (session_dir / "Mini project").mkdir(parents=True, exist_ok=True)
            continue
        elif is_practice:
            # Ràng buộc cho bài thực hành: chỉ có thư mục Bài tập
            (session_dir / "Bài tập").mkdir(parents=True, exist_ok=True)
            continue

        if session["lessons"]:
            for lesson in session["lessons"]:
                lesson_id = lesson["lesson_id"]
                lesson_title = lesson["title"]
                
                lesson_dir = get_or_rename_sanitized_folder(session_dir, lesson_id, f"{lesson_id} - {lesson_title}")
                lesson_dir.mkdir(parents=True, exist_ok=True)

                is_lesson_project_or_hackathon = any(kw in lesson_title.lower() or kw in session_title.lower() for kw in ["hackathon", "project", "đồ án", "dự án", "mini project"])
                is_lesson_practice = "thực hành" in lesson_title.lower() or "thực hành" in session_title.lower()

                if is_lesson_project_or_hackathon:
                    (lesson_dir / "Bài kiểm tra đầu giờ").mkdir(parents=True, exist_ok=True)
                    (lesson_dir / "Tài liệu đặc tả SRS").mkdir(parents=True, exist_ok=True)
                    (lesson_dir / "Mini project").mkdir(parents=True, exist_ok=True)
                    continue
                elif is_lesson_practice:
                    (lesson_dir / "Bài tập").mkdir(parents=True, exist_ok=True)
                    continue
                
                # Khởi tạo các thư mục tài nguyên con rỗng cho các lesson thông thường
                if "html" in requested_parts:
                    sub = lesson_dir / "Bài đọc"
                    sub.mkdir(parents=True, exist_ok=True)
                    with open(sub / "reading.html", "w", encoding="utf-8") as f:
                        f.write(f"<!-- Empty outline for {session_id} - {lesson_id}: {lesson_title} -->\n")
                
                if "slide" in requested_parts:
                    sub = lesson_dir / "Bài giảng"
                    sub.mkdir(parents=True, exist_ok=True)
                    with open(sub / "slides.md", "w", encoding="utf-8") as f:
                        f.write(f"<!-- Empty slide outline for {session_id} - {lesson_id}: {lesson_title} -->\n")
                
                if "quiz" in requested_parts:
                    sub = lesson_dir / "Câu hỏi Quizz"
                    sub.mkdir(parents=True, exist_ok=True)
                    with open(sub / "quiz.json", "w", encoding="utf-8") as f:
                        f.write("{}\n")
                
                if "video" in requested_parts or "video_script" in requested_parts:
                    sub = lesson_dir / "Video"
                    sub.mkdir(parents=True, exist_ok=True)
                    with open(sub / "SCRIPT.md", "w", encoding="utf-8") as f:
                        f.write(f"<!-- Empty video script outline for {session_id} - {lesson_id}: {lesson_title} -->\n")
                
                if "mindmap" in requested_parts:
                    sub = lesson_dir / "Mindmap"
                    sub.mkdir(parents=True, exist_ok=True)
                    with open(sub / "mindmap.md", "w", encoding="utf-8") as f:
                        f.write(f"<!-- Empty mindmap outline for {session_id} - {lesson_id}: {lesson_title} -->\n")
        else:
            # Session-level empty resources (practice/lab session - usually handled above but kept as fallback)
            if "html" in requested_parts:
                sub = session_dir / "Bài đọc"
                sub.mkdir(parents=True, exist_ok=True)
                with open(sub / "reading.html", "w", encoding="utf-8") as f:
                    f.write(f"<!-- Empty outline for {session_id} -->\n")
            if "slide" in requested_parts:
                sub = session_dir / "Bài giảng"
                sub.mkdir(parents=True, exist_ok=True)
                with open(sub / "slides.md", "w", encoding="utf-8") as f:
                    f.write(f"<!-- Empty slide outline for {session_id} -->\n")
            if "quiz" in requested_parts:
                sub = session_dir / "Câu hỏi Quizz"
                sub.mkdir(parents=True, exist_ok=True)
                with open(sub / "quiz.json", "w", encoding="utf-8") as f:
                    f.write("{}\n")
            if "video" in requested_parts or "video_script" in requested_parts:
                sub = session_dir / "Video"
                sub.mkdir(parents=True, exist_ok=True)
                with open(sub / "SCRIPT.md", "w", encoding="utf-8") as f:
                    f.write(f"<!-- Empty video script outline for {session_id} -->\n")
            if "mindmap" in requested_parts:
                sub = session_dir / "Mindmap"
                sub.mkdir(parents=True, exist_ok=True)
                with open(sub / "mindmap.md", "w", encoding="utf-8") as f:
                    f.write(f"<!-- Empty mindmap outline for {session_id} -->\n")

def project_structure_reviewer_agent(sessions, course_dir: Path, requested_parts: list, requested_session: str):
    print("\n=====================================================================")
    print(">>> GIAI ĐOẠN 1.2: AGENT QUÉT & THẨM ĐỊNH CẤU TRÚC KHUNG HỌC LIỆU <<<")
    print("=====================================================================")
    missing_elements = []
    for session in sessions:
        session_id = session["session_id"]
        session_title = session.get("title", "")
        req_sess_list = [r.strip() for r in requested_session.split(",")] if requested_session != "all" else ["all"]
        if "all" not in req_sess_list and not any(rs in session_id.lower() for rs in req_sess_list):
            continue

        session_dir = get_or_rename_sanitized_folder(course_dir, session_id, f"{session_id} - {session_title}")
        if not session_dir.exists():
            missing_elements.append(f"Thiếu thư mục session: {session_id}")
            continue

        is_project_or_hackathon = any(kw in session_title.lower() for kw in ["hackathon", "project", "đồ án", "dự án", "mini project"])
        is_practice = "thực hành" in session_title.lower()

        if is_project_or_hackathon:
            if not (session_dir / "Bài kiểm tra đầu giờ").exists():
                missing_elements.append(f"Thiếu thư mục 'Bài kiểm tra đầu giờ' tại {session_id}")
            if not (session_dir / "Tài liệu đặc tả SRS").exists():
                missing_elements.append(f"Thiếu thư mục 'Tài liệu đặc tả SRS' tại {session_id}")
            if not (session_dir / "Mini project").exists():
                missing_elements.append(f"Thiếu thư mục 'Mini project' tại {session_id}")
            continue
        elif is_practice:
            if not (session_dir / "Bài tập").exists():
                missing_elements.append(f"Thiếu thư mục 'Bài tập' tại {session_id}")
            continue

        if session["lessons"]:
            for lesson in session["lessons"]:
                lesson_id = lesson["lesson_id"]
                lesson_title = lesson["title"]
                lesson_dir = get_or_rename_sanitized_folder(session_dir, lesson_id, f"{lesson_id} - {lesson_title}")
                if not lesson_dir.exists():
                    missing_elements.append(f"Thiếu thư mục lesson: {session_id} -> {lesson_id}")
                    continue

                is_lesson_project_or_hackathon = any(kw in lesson_title.lower() or kw in session_title.lower() for kw in ["hackathon", "project", "đồ án", "dự án", "mini project"])
                is_lesson_practice = "thực hành" in lesson_title.lower() or "thực hành" in session_title.lower()

                if is_lesson_project_or_hackathon:
                    if not (lesson_dir / "Bài kiểm tra đầu giờ").exists():
                        missing_elements.append(f"Thiếu thư mục 'Bài kiểm tra đầu giờ' tại {session_id} -> {lesson_id}")
                    if not (lesson_dir / "Tài liệu đặc tả SRS").exists():
                        missing_elements.append(f"Thiếu thư mục 'Tài liệu đặc tả SRS' tại {session_id} -> {lesson_id}")
                    if not (lesson_dir / "Mini project").exists():
                        missing_elements.append(f"Thiếu thư mục 'Mini project' tại {session_id} -> {lesson_id}")
                    continue
                elif is_lesson_practice:
                    if not (lesson_dir / "Bài tập").exists():
                        missing_elements.append(f"Thiếu thư mục 'Bài tập' tại {session_id} -> {lesson_id}")
                    continue

                if "html" in requested_parts and not (lesson_dir / "Bài đọc" / "reading.html").exists():
                    missing_elements.append(f"Thiếu file reading.html tại {session_id} -> {lesson_id}")
                if "slide" in requested_parts and not (lesson_dir / "Bài giảng" / "slides.md").exists():
                    missing_elements.append(f"Thiếu file slides.md tại {session_id} -> {lesson_id}")
                if "quiz" in requested_parts and not (lesson_dir / "Câu hỏi Quizz" / "quiz.json").exists():
                    missing_elements.append(f"Thiếu file quiz.json tại {session_id} -> {lesson_id}")
                if ("video" in requested_parts or "video_script" in requested_parts) and not (lesson_dir / "Video" / "SCRIPT.md").exists():
                    missing_elements.append(f"Thiếu file SCRIPT.md tại {session_id} -> {lesson_id}")
                if "mindmap" in requested_parts and not (lesson_dir / "Mindmap" / "mindmap.md").exists():
                    missing_elements.append(f"Thiếu file mindmap.md tại {session_id} -> {lesson_id}")
        else:
            if "html" in requested_parts and not (session_dir / "Bài đọc" / "reading.html").exists():
                missing_elements.append(f"Thiếu file reading.html tại {session_id}")
            if "slide" in requested_parts and not (session_dir / "Bài giảng" / "slides.md").exists():
                missing_elements.append(f"Thiếu file slides.md tại {session_id}")
            if "quiz" in requested_parts and not (session_dir / "Câu hỏi Quizz" / "quiz.json").exists():
                missing_elements.append(f"Thiếu file quiz.json tại {session_id}")
            if ("video" in requested_parts or "video_script" in requested_parts) and not (session_dir / "Video" / "SCRIPT.md").exists():
                missing_elements.append(f"Thiếu file SCRIPT.md tại {session_id}")
            if "mindmap" in requested_parts and not (session_dir / "Mindmap" / "mindmap.md").exists():
                missing_elements.append(f"Thiếu file mindmap.md tại {session_id}")

    report_path = course_dir / "structure_review_report.md"
    status = "APPROVED" if not missing_elements else "REJECTED"
    
    report_content = f"""# 📂 BÁO CÁO DUYỆT KHUNG CẤU TRÚC DỰ ÁN (PROJECT STRUCTURE REVIEW)
**Đường dẫn dự án:** `{course_dir}`

## 📊 Kết quả Thẩm định
* **Trạng thái:** `{status}`
* **Thời gian quét:** 2026-07-16

## 🔍 Chi tiết đánh giá
{"Mọi thư mục Session và Lesson rỗng đã được tạo lập thành công và đầy đủ cấu trúc khung rỗng." if not missing_elements else "Phát hiện các lỗi cấu trúc sau:"}

{chr(10).join([f"* ❌ {err}" for err in missing_elements]) if missing_elements else "* ✅ Cấu trúc thư mục đạt chuẩn outline ban đầu.*"}
"""
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_content)
    print(f"  [Structure Reviewer] Đã lưu báo cáo duyệt cấu trúc tại: {report_path.resolve()}")

    if missing_elements:
        raise ValueError(f"Duyệt cấu trúc thất bại. Xem chi tiết tại {report_path.name}")
    else:
        print("  [Structure Reviewer] ĐÃ PHÊ DUYỆT cấu trúc dự án. Tiếp tục giai đoạn xây dựng nội dung học liệu...")

def verify_previous_lessons_completed(sessions, current_session_id: str, current_lesson_id: str, course_dir: Path, requested_parts: list):
    # Flatten all lessons across sessions in sequence
    ordered_lessons = []
    for session in sessions:
        s_id = session["session_id"]
        s_title = session["title"]
        if session["lessons"]:
            for lesson in session["lessons"]:
                l_id = lesson["lesson_id"]
                l_title = lesson["title"]
                ordered_lessons.append((s_id, s_title, l_id, l_title))
        else:
            ordered_lessons.append((s_id, s_title, "", ""))
            
    # Find index of current
    current_idx = -1
    for idx, (s_id, _, l_id, _) in enumerate(ordered_lessons):
        if s_id == current_session_id and (l_id == current_lesson_id or (not l_id and not current_lesson_id)):
            current_idx = idx
            break
            
    if current_idx <= 0:
        return
        
    # Check all lessons before current_idx
    for idx in range(current_idx):
        prev_session_id, prev_session_title, prev_lesson_id, prev_lesson_title = ordered_lessons[idx]
        
        is_project_or_hackathon = any(kw in prev_session_title.lower() or kw in prev_lesson_title.lower() for kw in ["hackathon", "project", "đồ án", "dự án", "mini project"])
        is_practice = "thực hành" in prev_session_title.lower() or "thực hành" in prev_lesson_title.lower()
        if is_project_or_hackathon or is_practice:
            continue
            
        prev_session_dir = get_or_rename_sanitized_folder(course_dir, prev_session_id, f"{prev_session_id} - {prev_session_title}")
        
        if prev_lesson_id:
            prev_dir = get_or_rename_sanitized_folder(prev_session_dir, prev_lesson_id, f"{prev_lesson_id} - {prev_lesson_title}")
            name_str = f"{prev_session_id} -> {prev_lesson_id}"
        else:
            prev_dir = prev_session_dir
            name_str = prev_session_id
            
        uncompleted = []
        if "html" in requested_parts:
            html_file = prev_dir / "Bài đọc" / "reading.html"
            if not html_file.exists() or html_file.stat().st_size < 500 or "<!-- Empty outline" in html_file.read_text(encoding="utf-8"):
                uncompleted.append("Bài đọc/reading.html chưa được sinh nội dung chi tiết hoặc quá ngắn.")
        if "slide" in requested_parts:
            slide_file = prev_dir / "Bài giảng" / "slides.md"
            if not slide_file.exists() or slide_file.stat().st_size < 500 or "<!-- Empty slide outline" in slide_file.read_text(encoding="utf-8"):
                uncompleted.append("Bài giảng/slides.md chưa được sinh nội dung chi tiết hoặc quá ngắn.")
        if "quiz" in requested_parts:
            quiz_file = prev_dir / "Câu hỏi Quizz" / "quiz.json"
            if not quiz_file.exists() or quiz_file.stat().st_size < 20:
                uncompleted.append("Câu hỏi Quizz/quiz.json chưa được sinh nội dung chi tiết.")
        if "video" in requested_parts or "video_script" in requested_parts:
            video_file = prev_dir / "Video" / "SCRIPT.md"
            if not video_file.exists() or video_file.stat().st_size < 500 or "<!-- Empty video script outline" in video_file.read_text(encoding="utf-8"):
                uncompleted.append("Video/SCRIPT.md chưa được sinh nội dung chi tiết hoặc quá ngắn.")
        if "mindmap" in requested_parts:
            mindmap_file = prev_dir / "Mindmap" / "mindmap.md"
            if not mindmap_file.exists() or mindmap_file.stat().st_size < 500 or "<!-- Empty mindmap outline" in mindmap_file.read_text(encoding="utf-8"):
                uncompleted.append("Mindmap/mindmap.md chưa được sinh nội dung chi tiết hoặc quá ngắn.")
                
        if uncompleted:
            feedback_details = "\n".join([f"  - {item}" for item in uncompleted])
            raise ValueError(
                f"\n[RÀNG BUỘC TUẦN TỰ NGHIÊM NGẶT] Bắt buộc hoàn thành và phê duyệt học liệu trước đó:\n"
                f"📌 {name_str} chưa được sinh hoặc chưa đạt chuẩn phê duyệt:\n"
                f"{feedback_details}\n"
                f"Yêu cầu: AI phải tuân thủ kỷ luật sư phạm, thà làm chậm và chất lượng còn hơn làm ẩu làm nhanh!"
            )

def main():
    print("=====================================================================")
    print("Starting Multi-Agent Learning Content Factory (Antigravity Workflow)")
    print("=====================================================================")

    import argparse
    parser = argparse.ArgumentParser(description="Multi-Agent Learning Content Factory")
    parser.add_argument("--pm", type=str, default=r"d:\Rikkei Education\Elearning_Agent\Learning-Material\pms\PM_Web_Application_With_FastAPI.xlsx", help="Path to PM Excel sheet")
    parser.add_argument("--session", type=str, default="all", help="Session ID to run (e.g. 'Session 01', 'Session 02', or 'all')")
    parser.add_argument("--parts", type=str, default="all", help="Comma-separated parts to generate (html,slide,quiz,video,mindmap or all)")
    parser.add_argument("--force", action="store_true", help="Force rebuild artifacts even if already approved in checkpoint")
    parser.add_argument("--approve-pm", action="store_true", help="Approve the PM program and bypass the PM review blocker")
    parser.add_argument("--obsidian", action="store_true", help="Generate Obsidian Vault for the entire course structure")
    parser.add_argument("--scorm", action="store_true", help="Export compiled lessons to SCORM 1.2 .zip package for LMS import (Moodle, Canvas, etc.)")
    parser.add_argument("--cache-stats", action="store_true", help="Show Semantic Cache statistics and exit")
    args = parser.parse_args()

    excel_path = args.pm
    if not os.path.exists(excel_path):
        print(f"Error: Excel PM file not found at {excel_path}")
        return

    # ── Cache Stats (chạy độc lập, không cần PM) ──
    if args.cache_stats:
        try:
            from core.semantic_cache import get_cache_stats, cache_invalidate_old
            cache_invalidate_old()
            stats = get_cache_stats()
            print("\n====== 📊 Semantic Cache Statistics ======")
            print(f"  Tổng response đã cache : {stats['total_cached_responses']}")
            print(f"  Tổng lần cache HIT     : {stats['total_cache_hits']}")
            print(f"  Ước tính tokens tiết kiệm: ~{stats['estimated_tokens_saved']:,}")
            print("  Phân tích theo Agent:")
            for agent, info in stats.get("by_agent", {}).items():
                print(f"    {agent}: {info['cached']} cached, {info['hits']} hits")
            print("==========================================\n")
        except Exception as e:
            print(f"[Cache Stats Error] {e}")
        return

    # ── SCORM Export (chạy sau khi đã compile output/) ──
    if args.scorm:
        course_dir_name = Path(excel_path).stem.strip().replace(" ", "_").replace("-", "_")
        output_course_dir = os.path.join("output", course_dir_name)
        if not os.path.exists(output_course_dir):
            print(f"[SCORM] Lỗi: Thư mục output chưa tồn tại: {output_course_dir}")
            print("Hãy chạy pipeline biên dịch học liệu trước: python main.py --pm ... --approve-pm")
            return
        try:
            from core.scorm_exporter import export_scorm_package
            export_scorm_package(
                output_course_dir=output_course_dir,
                course_name=course_dir_name.replace("_", " "),
            )
        except Exception as e:
            print(f"[SCORM Export Error] {e}")
        return

    if args.obsidian:
        sessions = parse_all_sessions(excel_path)
        print(f"Loaded {len(sessions)} sessions for Obsidian Knowledge Vault generation.")

        # Chạy Prerequisite Guard trước để có dependency graph
        prerequisite_data = None
        try:
            from agents.prerequisite_guard_agent import run_prerequisite_check_for_pm
            tech_stack_for_obsidian = Path(excel_path).stem.lower()
            stack = "python/fastapi"
            if "nestjs" in tech_stack_for_obsidian or "nest" in tech_stack_for_obsidian:
                stack = "typescript/nestjs"
            elif "core" in tech_stack_for_obsidian or "basic" in tech_stack_for_obsidian:
                stack = "python/core"
            _, prerequisite_data = run_prerequisite_check_for_pm(sessions, stack)
            print(f"  [Obsidian] Prerequisite analysis complete: "
                  f"{prerequisite_data.get('stats', {}).get('total_violations', 0)} violations found.")
        except Exception as e:
            print(f"  [Obsidian] Prerequisite check skipped: {e}")

        # Dùng ObsidianKnowledgeLinker (mới) thay vì obsidian_exporter (cũ)
        try:
            from core.obsidian_knowledge_linker import generate_knowledge_vault
            vault_path = generate_knowledge_vault(
                excel_path=excel_path,
                sessions=sessions,
                prerequisite_data=prerequisite_data,
            )
            print(f"\n✅ Knowledge Vault đã tạo thành công: {vault_path}")
        except ImportError:
            # Fallback về exporter cũ nếu linker chưa được cài
            from core.obsidian_exporter import generate_obsidian_vault
            generate_obsidian_vault(excel_path)
        return

    requested_parts = [p.strip().lower() for p in args.parts.split(",")] if args.parts != "all" else ["html", "slide", "quiz", "video", "mindmap"]
    requested_sessions = [s.strip().lower() for s in args.session.split(",")] if args.session != "all" else ["all"]

    print(f"Loading spreadsheet: {excel_path}")
    sessions = parse_all_sessions(excel_path)
    print(f"Successfully loaded {len(sessions)} sessions from spreadsheet.")
    if "all" not in requested_sessions:
        sessions = [s for s in sessions if any(rs in s["session_id"].lower() for rs in requested_sessions)]
        print(f"Filtered to {len(sessions)} sessions matching {requested_sessions}.")

    # Compile the workflow graph
    workflow = compile_learning_content_workflow()

    # Create output base directory
    output_base_dir = Path("output")
    output_base_dir.mkdir(exist_ok=True)

    # Get course directory name dynamically from the Excel filename
    course_dir_name = Path(excel_path).stem.strip().replace(" ", "_").replace("-", "_")
    course_dir = output_base_dir / course_dir_name
    course_dir.mkdir(parents=True, exist_ok=True)

    # Detect technology stack dynamically
    tech_stack = "python/fastapi"
    fn = os.path.basename(excel_path).lower()
    sn = course_dir_name.lower()
    if "nestjs" in fn or "nestjs" in sn or "nest" in fn or "nest" in sn:
        tech_stack = "typescript/nestjs"
    elif "react" in fn or "react" in sn:
        tech_stack = "typescript/react"
    elif "java" in fn or "java" in sn or "springboot" in fn or "springboot" in sn:
        tech_stack = "java/springboot"
    elif "core" in fn or "core" in sn or "basic" in fn or "basic" in sn:
        tech_stack = "python/core"
    print(f"Detected course technology stack: {tech_stack}")

    if not args.approve_pm:
        print("\n=========================================")
        print(">>> Giai đoạn 0: Thẩm định PM Input <<<")
        print("=========================================")
        from agents.reviewer_agents import pm_reviewer_agent
        full_curriculum_json = json.dumps(sessions, ensure_ascii=False)
        report = pm_reviewer_agent(full_curriculum_json, tech_stack)
        
        report_path = course_dir / "pm_review_report.md"
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(report)
        print(f"\n[CHỜ DUYỆT PM] Hệ thống đã đánh giá file PM và xuất báo cáo tại {report_path}.")
        
        choice = input("Bạn có muốn AI tự động chỉnh sửa file Excel PM dựa trên các đề xuất này không? (y/n): ")
        if choice.strip().lower() == 'y':
            from agents.reviewer_agents import pm_updater_agent
            print("Đang tiến hành cập nhật PM...")
            updated_json = pm_updater_agent(full_curriculum_json, report, tech_stack)
            
            from core.pm_updater_excel import export_updated_pm_to_excel
            import json as json_lib
            try:
                # Agent might return markdown wrapped json, strip it
                clean_json = updated_json.replace("```json", "").replace("```", "").strip()
                parsed_json = json_lib.loads(clean_json)
                new_excel_path = str(Path(excel_path).with_name(f"{Path(excel_path).stem}_AI_Updated.xlsx"))
                export_updated_pm_to_excel(parsed_json, excel_path, new_excel_path)
                print(f"Đã lưu file PM mới tại: {new_excel_path}")
                print("Vui lòng chạy lại hệ thống với file Excel mới (--pm) và cờ --approve-pm!")
            except Exception as e:
                print(f"Lỗi khi cập nhật file Excel: {e}")
            return
        else:
            print("Vui lòng xem báo cáo, tự điều chỉnh file PM và chạy lại với cờ --approve-pm!")
            return

    # Khởi tạo và duyệt cấu trúc thư mục/file rỗng trước khi chạy sinh nội dung
    initialize_skeleton_structure(sessions, course_dir, requested_parts, args.session.strip().lower())
    project_structure_reviewer_agent(sessions, course_dir, requested_parts, args.session.strip().lower())

    summary = []

    for session in sessions:
        session_id = session["session_id"]
        session_title = session.get("title", "")

        if "all" not in requested_sessions and not any(rs in session_id.lower() for rs in requested_sessions):
            continue

        print(f"\n=====================================================================")
        print(f"PROCESSING SESSION: {session_id} - {session_title}")
        print(f"=====================================================================")

        is_project_or_hackathon = any(kw in session_title.lower() for kw in ["hackathon", "project", "đồ án", "dự án", "mini project"])
        is_practice = "thực hành" in session_title.lower()

        if is_project_or_hackathon or is_practice:
            print(f"\n  ---> Processing Session-Level: {session_id} - {session_title}")
            # verify_previous_lessons_completed(sessions, session_id, "", course_dir, requested_parts)
            
            if args.approve_pm:
                session_dir = get_or_rename_sanitized_folder(course_dir, session_id, f"{session_id} - {session_title}")
                session_dir.mkdir(parents=True, exist_ok=True)
                
                # Collect all previous lessons text
                previous_lessons_text = ""
                for s in sessions:
                    if s["session_id"] == session_id:
                        break
                    for l in s.get("lessons", []):
                        previous_lessons_text += f"- {l['lesson_id']}: {l['title']} ({l.get('details', '')})\n"
                
                if is_project_or_hackathon:
                    print(f"  [Project/Hackathon] Generating Mini Project templates via Agent for: {session_title}")
                    from agents.project_agents import generate_mini_project_session
                    generate_mini_project_session(
                        session_id=session_id,
                        session_title=session_title,
                        session_dir_path=str(session_dir),
                        tech_stack=tech_stack,
                        previous_lessons_text=previous_lessons_text
                    )
                elif is_practice:
                    print(f"  [Practice] Generating practice exercises via Agent for: {session_title}")
                    from agents.practice_agents import generate_practice_session_exercises
                    generate_practice_session_exercises(
                        session_id=session_id,
                        session_title=session_title,
                        session_dir_path=str(session_dir),
                        tech_stack=tech_stack,
                        previous_lessons_text=previous_lessons_text
                    )
                
                summary.append({
                    "session_id": session_id,
                    "lesson_id": "",
                    "title": session_title,
                    "html_file": "Generated Project Templates" if is_project_or_hackathon else "Generated Practice Assignments",
                    "slides_file": "Skipped (Project/Practice)",
                    "quiz_file": "Skipped (Project/Practice)",
                    "video_script_file": "Skipped (Project/Practice)",
                    "mindmap_file": "Skipped (Project/Practice)",
                    "status": "APPROVED",
                    "review_count": 0
                })
                continue

        # If session has lessons, loop and process each lesson
        if session["lessons"]:
            previous_lessons = []
            for idx, lesson in enumerate(session["lessons"]):
                lesson_id = lesson["lesson_id"]
                lesson_title = lesson["title"]
                lesson_details = lesson["details"]
                expected_output = lesson["expected_output"]
                

                print(f"\n  ---> Processing: {session_id} - {lesson_id}: {lesson_title}")
                
                # Kiểm tra ràng buộc tuần tự nghiêm ngặt đối với các bài học trước đó
                # verify_previous_lessons_completed(sessions, session_id, lesson_id, course_dir, requested_parts)
                
                # Initialize state for this specific lesson
                state: AgentState = {
                    "session_id": session_id,
                    "lesson_id": lesson_id,
                    "pm_input": json.dumps(lesson, ensure_ascii=False),
                    "full_curriculum": json.dumps(get_context_curriculum(sessions, session_id), ensure_ascii=False),
                    "time_reference": {
                        "weeks": 1,
                        "hours_per_week": 10
                    },
                    "learning_outcomes": {},
                    "program_structure": {},
                    "core_ssot": {
                        "session_title": lesson_title,
                        "lesson_details": lesson_details,
                        "expected_output": expected_output
                    },
                    "previous_lessons": previous_lessons.copy(),
                    "artifacts_status": {
                        "html": "Pending",
                        "slide": "Pending",
                        "quiz": "Pending",
                        "video_script": "Pending",
                        "mindmap": "Pending",
                        "session": "Pending"
                    },
                    "course_dir_name": course_dir_name,
                    "technology_stack": tech_stack,
                    "html_content": "",
                    "slide_markdown": "",
                    "quiz_json": {},
                    "video_script_markdown": "",
                    "mindmap_markdown": "",
                    "review_logs": [],
                    "requested_parts": requested_parts,
                    "force_rebuild": args.force,
                    "pm_approved": args.approve_pm
                }

                is_project_or_hackathon = any(kw in session_title.lower() or kw in lesson_title.lower() for kw in ["hackathon", "project", "đồ án", "dự án", "mini project"])
                is_practice = "thực hành" in session_title.lower() or "thực hành" in lesson_title.lower()
                
                if args.approve_pm and (is_project_or_hackathon or is_practice):
                    session_dir = get_or_rename_sanitized_folder(course_dir, session_id, f"{session_id} - {session_title}")
                    lesson_dir = get_or_rename_sanitized_folder(session_dir, lesson_id, f"{lesson_id} - {lesson_title}")
                    lesson_dir.mkdir(parents=True, exist_ok=True)
                    
                    if is_project_or_hackathon:
                        print(f"  [Skip] Skipping AI generation for Project/Hackathon: {lesson_title}")
                        (lesson_dir / "Bài kiểm tra đầu giờ").mkdir(exist_ok=True)
                        (lesson_dir / "Tài liệu đặc tả SRS").mkdir(exist_ok=True)
                        (lesson_dir / "Mini project").mkdir(exist_ok=True)
                    elif is_practice:
                        print(f"  [Skip] Skipping AI generation for Practice Session: {lesson_title}")
                        (lesson_dir / "Bài tập").mkdir(exist_ok=True)
                    
                    # Update summary and skip the graph
                    summary.append({
                        "session_id": session_id,
                        "lesson_id": lesson_id,
                        "title": lesson_title,
                        "html_file": "Skipped (Project/Practice)",
                        "slides_file": "Skipped (Project/Practice)",
                        "quiz_file": "Skipped (Project/Practice)",
                        "video_script_file": "Skipped (Project/Practice)",
                        "mindmap_file": "Skipped (Project/Practice)",
                        "status": "APPROVED",
                        "review_count": 0
                    })
                    continue

                # Run the workflow graph for this lesson
                final_state = workflow.run(state)
                
                # Append to previous_lessons for next iteration
                previous_lessons.append({
                    "lesson_id": lesson_id,
                    "title": lesson_title,
                    "details": lesson_details,
                    "expected_output": expected_output
                })

                # Write output files to output/<course_name>/session_xx/lesson_yy/
                session_dir = get_or_rename_sanitized_folder(course_dir, session_id, f"{session_id} - {session_title}")
                lesson_dir = get_or_rename_sanitized_folder(session_dir, lesson_id, f"{lesson_id} - {lesson_title}")
                lesson_dir.mkdir(parents=True, exist_ok=True)

                # Save HTML Reading to "Bài đọc" subfolder
                if "html" in requested_parts and final_state.get("html_content"):
                    html_sub = lesson_dir / "Bài đọc"
                    html_sub.mkdir(parents=True, exist_ok=True)
                    html_path = html_sub / "reading.html"
                    with open(html_path, "w", encoding="utf-8") as f:
                        f.write(final_state.get("html_content", ""))
                else:
                    html_path = "Skipped"

                # Save Self-Test Questions to "Câu hỏi Bài đọc" subfolder
                if "html" in requested_parts and final_state.get("self_test_markdown"):
                    selftest_sub = lesson_dir / "Câu hỏi Bài đọc"
                    selftest_sub.mkdir(parents=True, exist_ok=True)
                    selftest_path = selftest_sub / "self_test.md"
                    with open(selftest_path, "w", encoding="utf-8") as f:
                        f.write(final_state.get("self_test_markdown", ""))

                # Save Slides to "Bài giảng" subfolder
                if "slide" in requested_parts and final_state.get("slide_markdown"):
                    slide_sub = lesson_dir / "Bài giảng"
                    slide_sub.mkdir(parents=True, exist_ok=True)
                    slides_path = slide_sub / "slides.md"
                    with open(slides_path, "w", encoding="utf-8") as f:
                        f.write(final_state.get("slide_markdown", ""))
                else:
                    slides_path = "Skipped"

                # Save Quiz to "Câu hỏi Quizz" subfolder (Excel strictly following Format_Quizz_Lesson.xlsx + JSON)
                if "quiz" in requested_parts and final_state.get("quiz_json"):
                    quiz_sub = lesson_dir / "Câu hỏi Quizz"
                    quiz_sub.mkdir(parents=True, exist_ok=True)
                    quiz_path = quiz_sub / "quiz.json"
                    with open(quiz_path, "w", encoding="utf-8") as f:
                        json.dump(final_state.get("quiz_json", {}), f, indent=2, ensure_ascii=False)
                    
                    from core.quiz_excel import export_lesson_quiz_to_excel
                    import re
                    sanitized_lesson_title = re.sub(r'[\s\-\:\,\.\?\!\(\)]+', '_', lesson_title)
                    sanitized_lesson_title = re.sub(r'^_+|_+$', '', sanitized_lesson_title)
                    s_num_str = session_id.replace(" ", "")
                    l_num_str = lesson_id.replace(" ", "")
                    excel_filename = f"Quizz_{s_num_str}_{l_num_str}_{sanitized_lesson_title}.xlsx"
                    excel_path = quiz_sub / excel_filename
                    
                    quiz_items = final_state.get("quiz_json", {}).get("lesson_quiz") or final_state.get("quiz_json", {}).get("quiz") or []
                    export_lesson_quiz_to_excel(quiz_items, str(excel_path))
                    quiz_reported_path = str(excel_path)
                else:
                    quiz_reported_path = "Skipped"

                # Save Video Script and Scaffold Modular HyperFrames Project to "Video" subfolder
                if ("video" in requested_parts or "video_script" in requested_parts) and final_state.get("video_script_markdown"):
                    video_sub = lesson_dir / "Video"
                    video_sub.mkdir(parents=True, exist_ok=True)
                    video_script_path = video_sub / "SCRIPT.md"
                    with open(video_script_path, "w", encoding="utf-8") as f:
                        f.write(final_state.get("video_script_markdown", ""))
                        
                    v_json = final_state.get("video_script_json")
                    if v_json:
                        # 1. Run hyperframes_writer_agent to scaffold project with dynamic premium layouts
                        from agents.hyperframes_writer_agent import hyperframes_writer_agent
                        final_state = hyperframes_writer_agent(final_state)
                        
                        project_path_str = final_state.get("hyperframes_project_path")
                        if project_path_str:
                            video_project_dir = Path(project_path_str)
                            assets_dir = video_project_dir / "assets"
                            tts_dir = assets_dir / "tts"
                            
                            # 2. Link or copy the default assets intro.mp4, outro.mp4, bg-music.mp3
                            project_root = Path(__file__).resolve().parent
                            asset_search_paths = [
                                project_root / "hyperframes" / "dev-tutorial-video" / "courses" / "fundamental_python" / "assets",
                                project_root / "hyperframes" / "java-intro-video" / "courses" / "java_intro" / "assets",
                                project_root / "assets"
                            ]
                            
                            import shutil
                            for p in asset_search_paths:
                                if p.exists():
                                    try:
                                        for asset_name in ["intro.mp4", "outro.mp4", "bg-music.mp3"]:
                                            src_file = p / asset_name
                                            dest_file = assets_dir / asset_name
                                            if src_file.exists() and not dest_file.exists():
                                                try:
                                                    os.link(str(src_file), str(dest_file))
                                                except Exception:
                                                    shutil.copy2(src_file, dest_file)
                                    except Exception as e:
                                        print("  [Assets Error] Failed to link or copy assets:", e)
                            
                            # 3. Generate prepare_tts.py to run TTS for the scenes
                            scenes = v_json.get("scenes", [])
                            py_scenes = []
                            for idx, scene in enumerate(scenes):
                                s_id = scene.get("scene_id", f"Scene_{idx+1:02d}")
                                narration = scene.get("narration", "")
                                py_scenes.append({"id": s_id, "text": narration})
                            
                            py_content = f"""import os
import json
import subprocess
import sys
import io
from pathlib import Path
from gtts import gTTS

# Set console encoding to UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', write_through=True)

scenes = {json.dumps(py_scenes, indent=2, ensure_ascii=False)}

tts_dir = Path("assets/tts")
tts_dir.mkdir(parents=True, exist_ok=True)

all_durations = {{}}

for scene in scenes:
    scene_id = scene["id"]
    text = scene["text"]
    if not text:
        continue
        
    mp3_path = tts_dir / f"{{scene_id}}.mp3"
    print(f"Generating TTS for {{scene_id}} using gTTS...")
    try:
        tts = gTTS(text=text, lang="vi")
        if mp3_path.exists():
            mp3_path.unlink()
        tts.save(str(mp3_path))
    except Exception as e:
        print(f"Failed to generate TTS for {{scene_id}}: {{e}}.")
        if not mp3_path.exists():
            print("Generating fallback silent audio...")
            try:
                fallback_dur = max(5.0, len(text.split()) / 2.5)
                cmd = f'ffmpeg -y -f lavfi -i anullsrc=r=44100:cl=mono -t {{fallback_dur}} -q:a 2 "{{mp3_path}}"'
                subprocess.run(cmd, shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            except Exception as fe:
                print(f"Failed to generate fallback silent audio: {{fe}}")
                
    if mp3_path.exists():
        try:
            cmd = f'ffprobe -v error -show_entries format=duration -of default=noprint_wrappers=1:nokey=1 "{{mp3_path}}"'
            probe = subprocess.check_output(cmd, shell=True).decode("utf-8").strip()
            duration = float(probe)
            all_durations[scene_id] = round(duration, 2) + 0.5
        except Exception as e:
            all_durations[scene_id] = max(5.0, len(text.split()) / 2.5)
    else:
        all_durations[scene_id] = max(5.0, len(text.split()) / 2.5)

with open(tts_dir / "durations.json", "w", encoding="utf-8") as f:
    json.dump(all_durations, f, indent=2, ensure_ascii=False)
print("TTS Generation Complete!")
"""
                            py_path = video_project_dir / "prepare_tts.py"
                            with open(py_path, "w", encoding="utf-8") as f:
                                f.write(py_content)
                                
                            # 4. Execute python prepare_tts.py
                            print("  [TTS] Running prepare_tts.py in video project directory...")
                            import subprocess
                            try:
                                cwd_str = str(video_project_dir)
                                if cwd_str.startswith("\\\\?\\"):
                                    cwd_str = cwd_str[4:]
                                subprocess.run(["python", "prepare_tts.py"], cwd=cwd_str, check=True, shell=True)
                            except Exception as e:
                                print(f"  [TTS Error] Python script execution failed: {e}")
                                
                            # 5. Read back actual durations from durations.json
                            durations_file = tts_dir / "durations.json"
                            durations = {}
                            if durations_file.exists():
                                with open(durations_file, "r", encoding="utf-8") as f:
                                    durations = json.load(f)
                                    
                            # Update scene durations in the blueprint JSON
                            has_duration_changed = False
                            for scene in scenes:
                                s_id = scene.get("scene_id")
                                if s_id in durations:
                                    old_dur = scene.get("duration")
                                    new_dur = durations[s_id]
                                    if old_dur != new_dur:
                                        scene["duration"] = new_dur
                                        has_duration_changed = True
                                        
                            if has_duration_changed:
                                # Re-calculate start times and total duration
                                current_time = 9.24  # intro duration
                                for scene in scenes:
                                    scene["start_at_root"] = current_time
                                    current_time += scene["duration"]
                                v_json["total_duration"] = current_time + 12.15  # outro duration
                                
                                # Update markdown with the new durations (optional but good)
                                print("  [TTS Sync] Durations changed! Re-scaffolding the video composition...")
                                final_state = hyperframes_writer_agent(final_state)
                else:
                    video_script_path = "Skipped"

                # Save Mindmap MD to "Mindmap" subfolder
                if "mindmap" in requested_parts and final_state.get("mindmap_markdown"):
                    mindmap_sub = lesson_dir / "Mindmap"
                    mindmap_sub.mkdir(parents=True, exist_ok=True)
                    mindmap_path = mindmap_sub / "mindmap.md"
                    with open(mindmap_path, "w", encoding="utf-8") as f:
                        f.write(final_state.get("mindmap_markdown", ""))
                else:
                    mindmap_path = "Skipped"

                print(f"  [Publisher] Saved learning assets to {lesson_dir}/")
                
                summary.append({
                    "session_id": session_id,
                    "lesson_id": lesson_id,
                    "title": lesson_title,
                    "html_file": str(html_path),
                    "slides_file": str(slides_path),
                    "quiz_file": quiz_reported_path,
                    "video_script_file": str(video_script_path),
                    "mindmap_file": str(mindmap_path),
                    "status": final_state.get("artifacts_status", {}).get("session", "FAILED"),
                    "review_count": len(final_state.get("review_logs", []))
                })
            
            # Compile session-level unified HTML and Mindmap
            from core.session_compilers import compile_session_html, compile_session_mindmap
            s_dir = get_or_rename_sanitized_folder(course_dir, session_id, f"{session_id} - {session_title}")
            full_session_name = f"{session_id} - {session_title}"
            if "html" in requested_parts:
                compile_session_html(s_dir, full_session_name)
            if "mindmap" in requested_parts:
                compile_session_mindmap(s_dir, full_session_name)

            # Generate Session-Level Homework (Theory Session Exercises)
            if args.approve_pm:
                print(f"  [Homework] Generating theoretical session homework for: {session_title}")
                session_lessons_text = ""
                for idx, lesson in enumerate(session.get("lessons", [])):
                    session_lessons_text += f"- {lesson['lesson_id']}: {lesson['title']} ({lesson.get('details', '')})\n"
                
                from agents.homework_agents import generate_session_homework
                generate_session_homework(
                    session_id=session_id,
                    session_title=session_title,
                    session_dir_path=str(s_dir),
                    tech_stack=tech_stack,
                    previous_lessons_text=session_lessons_text
                )
        else:
            # Session-level fallback for sessions without sub-lessons (e.g. practical/lab session)
            print(f"\n  ---> Processing Session-Level practice: {session_id} - {session_title}")
            
            # Kiểm tra ràng buộc tuần tự nghiêm ngặt đối với các bài học trước đó
            verify_previous_lessons_completed(sessions, session_id, "", course_dir, requested_parts)
            
            # Initialize state for this specific session
            state: AgentState = {
                "session_id": session_id,
                "lesson_id": "",
                "pm_input": json.dumps(session, ensure_ascii=False),
                "full_curriculum": json.dumps(get_context_curriculum(sessions, session_id), ensure_ascii=False),
                "time_reference": {
                    "weeks": 1,
                    "hours_per_week": 10
                },
                "learning_outcomes": {},
                "program_structure": {},
                "core_ssot": {
                    "session_title": session_title,
                    "lesson_details": "",
                    "expected_output": ""
                },
                "artifacts_status": {
                    "html": "Pending",
                    "slide": "Pending",
                    "quiz": "Pending",
                    "video_script": "Pending",
                    "mindmap": "Pending",
                    "session": "Pending"
                },
                "course_dir_name": course_dir_name,
                "technology_stack": tech_stack,
                "html_content": "",
                "slide_markdown": "",
                "quiz_json": {},
                "video_script_markdown": "",
                "mindmap_markdown": "",
                "review_logs": [],
                "requested_parts": requested_parts,
                "force_rebuild": args.force,
                "pm_approved": args.approve_pm
            }

            is_project_or_hackathon = any(kw in session_title.lower() for kw in ["hackathon", "project", "đồ án", "dự án", "mini project"])
            is_practice = "thực hành" in session_title.lower()
            
            if args.approve_pm and (is_project_or_hackathon or is_practice):
                session_dir = course_dir / session_id.lower().replace(" ", "_")
                session_dir.mkdir(parents=True, exist_ok=True)
                
                if is_project_or_hackathon:
                    print(f"  [Skip] Skipping AI generation for Project/Hackathon: {session_title}")
                    (session_dir / "Tài liệu đặc tả (SRS)").mkdir(exist_ok=True)
                    (session_dir / "Mini Project").mkdir(exist_ok=True)
                    (session_dir / "Bài kiểm tra đầu giờ").mkdir(exist_ok=True)
                elif is_practice:
                    print(f"  [Skip] Skipping AI generation for Practice Session: {session_title}")
                    (session_dir / "Bài tập").mkdir(exist_ok=True)
                
                # Update summary and skip the graph
                summary.append({
                    "session_id": session_id,
                    "lesson_id": "",
                    "title": session_title,
                    "html_file": "Skipped (Project/Practice)",
                    "slides_file": "Skipped (Project/Practice)",
                    "quiz_file": "Skipped (Project/Practice)",
                    "video_script_file": "Skipped (Project/Practice)",
                    "mindmap_file": "Skipped (Project/Practice)",
                    "status": "APPROVED",
                    "review_count": 0
                })
                continue

            # Run the workflow graph for this session-level summary
            final_state = workflow.run(state)
            
            session_dir = get_or_rename_sanitized_folder(course_dir, session_id, f"{session_id} - {session_title}")
            session_dir.mkdir(parents=True, exist_ok=True)

            # Save HTML Reading to "Bài đọc" subfolder
            if "html" in requested_parts and final_state.get("html_content"):
                html_sub = session_dir / "Bài đọc"
                html_sub.mkdir(parents=True, exist_ok=True)
                html_path = html_sub / "reading.html"
                with open(html_path, "w", encoding="utf-8") as f:
                    f.write(final_state.get("html_content", ""))
            else:
                html_path = "Skipped"

            # Save Slides to "Bài giảng" subfolder
            if "slide" in requested_parts and final_state.get("slide_markdown"):
                slide_sub = session_dir / "Bài giảng"
                slide_sub.mkdir(parents=True, exist_ok=True)
                slides_path = slide_sub / "slides.md"
                with open(slides_path, "w", encoding="utf-8") as f:
                    f.write(final_state.get("slide_markdown", ""))
            else:
                slides_path = "Skipped"

            # Save Quiz to "Câu hỏi Quizz" subfolder
            if "quiz" in requested_parts and final_state.get("quiz_json"):
                quiz_sub = session_dir / "Câu hỏi Quizz"
                quiz_sub.mkdir(parents=True, exist_ok=True)
                quiz_path = quiz_sub / "quiz.json"
                with open(quiz_path, "w", encoding="utf-8") as f:
                    json.dump(final_state.get("quiz_json", {}), f, indent=2, ensure_ascii=False)
                from core.quiz_excel import export_lesson_quiz_to_excel
                s_num_str = session_id.replace(" ", "")
                excel_filename = f"Quizz_{s_num_str}_Thuc_hanh.xlsx"
                excel_path = quiz_sub / excel_filename
                quiz_items = final_state.get("quiz_json", {}).get("lesson_quiz") or final_state.get("quiz_json", {}).get("quiz") or []
                export_lesson_quiz_to_excel(quiz_items, str(excel_path))
                quiz_reported_path = str(excel_path)
            else:
                quiz_reported_path = "Skipped"

            # Save Video Script to "Kịch bản video" subfolder
            if ("video" in requested_parts or "video_script" in requested_parts) and final_state.get("video_script_markdown"):
                video_sub = session_dir / "Kịch bản video"
                video_sub.mkdir(parents=True, exist_ok=True)
                video_script_path = video_sub / "video_script.md"
                with open(video_script_path, "w", encoding="utf-8") as f:
                    f.write(final_state.get("video_script_markdown", ""))
            else:
                video_script_path = "Skipped"

            # Save Mindmap MD to "Mindmap" subfolder
            if "mindmap" in requested_parts and final_state.get("mindmap_markdown"):
                mindmap_sub = session_dir / "Mindmap"
                mindmap_sub.mkdir(parents=True, exist_ok=True)
                mindmap_path = mindmap_sub / "mindmap.md"
                with open(mindmap_path, "w", encoding="utf-8") as f:
                    f.write(final_state.get("mindmap_markdown", ""))
            else:
                mindmap_path = "Skipped"

            print(f"  [Publisher] Saved learning assets to {session_dir}/")
            
            summary.append({
                "session_id": session_id,
                "lesson_id": "",
                "title": session_title,
                "html_file": str(html_path),
                "slides_file": str(slides_path),
                "quiz_file": quiz_reported_path,
                "video_script_file": str(video_script_path),
                "mindmap_file": str(mindmap_path),
                "status": final_state.get("artifacts_status", {}).get("session", "FAILED"),
                "review_count": len(final_state.get("review_logs", []))
            })

            # Generate Session-Level Homework (Theory Session Exercises)
            if args.approve_pm and not (is_project_or_hackathon or is_practice):
                print(f"  [Homework] Generating theoretical session homework for: {session_title}")
                session_lessons_text = f"Kiến thức tổng quan của buổi học: {session_title}"
                if session.get("lessons"):
                    for idx, lesson in enumerate(session.get("lessons", [])):
                        session_lessons_text += f"\n- {lesson['lesson_id']}: {lesson['title']} ({lesson.get('details', '')})"
                
                from agents.homework_agents import generate_session_homework
                generate_session_homework(
                    session_id=session_id,
                    session_title=session_title,
                    session_dir_path=str(session_dir),
                    tech_stack=tech_stack,
                    previous_lessons_text=session_lessons_text
                )

        # Generate and save Session Quizzes (Entrance & Exit) in .xlsx format
        if "quiz" in requested_parts:
            print(f"  ---> Generating Session-Level Quizzes (Entrance & Exit) for {session_id}...")
            import re
            
            # Determine slide_title (first lesson's title if present, otherwise session title)
            slide_title = session["lessons"][0]["title"] if session["lessons"] else session_title
            
            # Sanitized titles for filename
            def get_sanitized_title(title: str) -> str:
                sanitized = re.sub(r'[\s\-\:\,\.\?\!\(\)]+', '_', title)
                sanitized = re.sub(r'^_+|_+$', '', sanitized)
                return sanitized
                
            session_num_str = session_id.replace(" ", "")
            sanitized_slide_title = get_sanitized_title(slide_title)
            
            # Paths
            session_title = next((s["title"] for s in sessions if s["session_id"] == session_id), "")
            session_dir = get_or_rename_sanitized_folder(course_dir, session_id, f"{session_id} - {session_title}")
            session_dir.mkdir(parents=True, exist_ok=True)
            
            entrance_path = session_dir / f"{session_num_str}._Quizz_Dau_Gio_{sanitized_slide_title}.xlsx"
            exit_path = session_dir / f"{session_num_str}._Quizz_Cuoi_Gio_{sanitized_slide_title}.xlsx"
            
            # Import generators
            from core.quiz_engine import generate_entrance_quiz, generate_exit_quiz
            from core.quiz_excel import export_quiz_to_excel
            from agents.creator_agents import get_base_topic_key
            
            # Current topic key
            current_topic = get_base_topic_key(session_id)
            
            # Previous topic key
            def get_previous_session_id(s_id: str) -> str:
                match = re.search(r'\d+', s_id)
                if match:
                    num = int(match.group(0))
                    if num > 1:
                        return f"Session {num-1:02d}"
                return "Session 01"
                
            previous_session_id = get_previous_session_id(session_id)
            previous_topic = get_base_topic_key(previous_session_id)
            
            # Generate quizzes
            entrance_qs = generate_entrance_quiz(session_id, current_topic, previous_topic, tech_stack)
            exit_qs = generate_exit_quiz(session_id, current_topic, tech_stack)
            
            # Export to Excel
            export_quiz_to_excel(entrance_qs, str(entrance_path))
            export_quiz_to_excel(exit_qs, str(exit_path))
            print(f"  [Publisher] Saved Entrance Quiz: {entrance_path.name}")
            print(f"  [Publisher] Saved Exit Quiz:     {exit_path.name}")

    print("\n=====================================================================")
    print("ALL SESSIONS COMPILED! FINAL GENERATION REPORT:")
    print("=====================================================================")
    for s in summary:
        lbl = f"{s['session_id']} - {s['lesson_id']}" if s['lesson_id'] else s['session_id']
        print(f"Unit: {lbl} - {s['title']}")
        print(f"  - Status: {s['status']}")
        print(f"  - Critique Reviews Attempted: {s['review_count']}")
        print(f"  - HTML Reading: {s['html_file']}")
        print(f"  - Slide Deck:   {s['slides_file']}")
        print(f"  - Quiz JSON:    {s['quiz_file']}")
        print(f"  - Video Script: {s['video_script_file']}")
        print(f"  - Mindmap MD:   {s['mindmap_file']}")
        print()

if __name__ == "__main__":
    main()