import os
import openpyxl

def dump_xlsx_to_txt(xlsx_path, txt_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(f"FILE: {os.path.basename(xlsx_path)}\n")
        for sheet in wb.sheetnames:
            f.write(f"\n=========================================\nSheet: {sheet}\n=========================================\n")
            ws = wb[sheet]
            for r_idx in range(1, ws.max_row + 1):
                row_vals = [cell.value for cell in ws[r_idx]]
                if any(val is not None for val in row_vals):
                    # Format cell values nicely
                    vals_str = []
                    for val in row_vals:
                        if val is None:
                            vals_str.append("")
                        else:
                            vals_str.append(str(val).replace('\n', ' ').strip())
                    # Write row
                    f.write(f"Row {r_idx:03d} | " + " | ".join(vals_str) + "\n")

if __name__ == "__main__":
    scratch_dir = r"d:\Rikkei Education\Elearning_Agent\Learning-Material\scratch"
    os.makedirs(scratch_dir, exist_ok=True)
    
    pm_path = r"d:\Rikkei Education\Elearning_Agent\Learning-Material\pms\PM_Python_Core_Synchronized.xlsx"
    pre_quiz = r"d:\Rikkei Education\Elearning_Agent\Learning-Material\output\PM_Python_Core_Synchronized\session_01\Session01._Quizz_Dau_Gio_Giới_thiệu_ngôn_ngữ_Python.xlsx"
    post_quiz = r"d:\Rikkei Education\Elearning_Agent\Learning-Material\output\PM_Python_Core_Synchronized\session_01\Session01._Quizz_Cuoi_Gio_Giới_thiệu_ngôn_ngữ_Python.xlsx"
    
    if os.path.exists(pm_path):
        dump_xlsx_to_txt(pm_path, os.path.join(scratch_dir, "pm_info.txt"))
    if os.path.exists(pre_quiz):
        dump_xlsx_to_txt(pre_quiz, os.path.join(scratch_dir, "pre_quiz_info.txt"))
    if os.path.exists(post_quiz):
        dump_xlsx_to_txt(post_quiz, os.path.join(scratch_dir, "post_quiz_info.txt"))
    print("Dumping finished successfully.")
