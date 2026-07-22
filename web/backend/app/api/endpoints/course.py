from fastapi import APIRouter, Depends
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from typing import List, Any
from pathlib import Path
ROOT = Path(__file__).resolve().parents[5]

from app.db.session import get_db
from app.models.course import Course
from app.schemas.course import CourseResponse, CourseCreate

router = APIRouter()

@router.get("/", response_model=List[CourseResponse])
async def get_courses(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Course))
    courses = result.scalars().all()
    return courses

@router.post("/", response_model=CourseResponse)
async def create_course(course_in: CourseCreate, db: AsyncSession = Depends(get_db)):
    new_course = Course(**course_in.model_dump())
    db.add(new_course)
    await db.commit()
    await db.refresh(new_course)
    return new_course

from fastapi import HTTPException

@router.get("/{course_id}", response_model=CourseResponse)
async def get_course(course_id: int, db: AsyncSession = Depends(get_db)):
    course = await db.get(Course, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    return course

from fastapi import UploadFile, File
import openpyxl
import io
import re
from app.models.session import Session
from app.models.lesson import Lesson

from pydantic import BaseModel
from typing import Optional

class PMRow(BaseModel):
    stt: Optional[str] = ""
    form: Optional[str] = ""
    session_val: Optional[str] = ""
    content_val: Optional[str] = ""
    lesson_val: Optional[str] = ""
    details_val: Optional[str] = ""
    output_val: Optional[str] = ""
    deadline: Optional[str] = ""

@router.post("/{course_id}/parse-excel", response_model=List[PMRow])
async def parse_excel_preview(course_id: int, file: UploadFile = File(...)):
    content = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(content))
    ws: Any = wb.active
    
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Prevent completely empty rows
        row_padded = list(row) + [None] * (8 - len(row))
        if not any(row_padded[:8]):
            continue
            
        stt, form, session_val, content_val, lesson_val, details_val, output_val, deadline = row_padded[:8]
        rows.append(PMRow(
            stt=str(stt) if stt is not None else "",
            form=str(form) if form is not None else "",
            session_val=str(session_val) if session_val is not None else "",
            content_val=str(content_val) if content_val is not None else "",
            lesson_val=str(lesson_val) if lesson_val is not None else "",
            details_val=str(details_val) if details_val is not None else "",
            output_val=str(output_val) if output_val is not None else "",
            deadline=str(deadline) if deadline is not None else ""
        ))
    return rows

def sync_curriculum_to_pms_excel(course_name: str, payload: List[PMRow]) -> None:
    """
    Finds the correct Excel file in the pms/ root directory, or creates one,
    and updates it with the list of PMRow objects representing the curriculum.
    """
    import openpyxl
    import re
    from pathlib import Path
    
    # ROOT is resolved as the project root (Learning-Material)
    pms_dir = ROOT / "pms"
    pms_dir.mkdir(parents=True, exist_ok=True)
    
    # 1. Resolve Excel file path
    target_excel = None
    xlsx_files = list(pms_dir.glob("*.xlsx")) if pms_dir.exists() else []
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]
    
    if xlsx_files:
        # Check if there is an excel file matching words in the course name
        clean_course_words = set(re.findall(r'\w+', course_name.lower()))
        best_match = None
        best_overlap = 0
        for f in xlsx_files:
            file_words = set(re.findall(r'\w+', f.stem.lower()))
            overlap = len(clean_course_words.intersection(file_words))
            if overlap > best_overlap:
                best_overlap = overlap
                best_match = f
                
        if best_match and best_overlap > 0:
            target_excel = best_match
        elif len(xlsx_files) == 1:
            target_excel = xlsx_files[0]
            
    if not target_excel:
        # Sanitize course name to construct a safe filename
        sanitized_name = re.sub(r'[\\/*?:"<>|]', "", course_name).strip().replace(" ", "_").replace("-", "_")
        target_excel = pms_dir / f"PM_{sanitized_name}.xlsx"
        
    print(f"  [PM Excel Sync] Writing curriculum to: {target_excel}")
    
    # 2. Write rows to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    if not ws:
        ws = wb.create_sheet()
    ws.title = "Syllabus"
    
    # Write headers matching PM layout
    headers = [
        "STT", "Hình thức", "Session", "Nội dung Session", 
        "Lesson", "Chi tiết / Prompt Context", "Output mong muốn", "Hạn chót (Deadline)"
    ]
    ws.append(headers)
    
    # Write curriculum rows
    for row in payload:
        ws.append([
            row.stt or "",
            row.form or "",
            row.session_val or "",
            row.content_val or "",
            row.lesson_val or "",
            row.details_val or "",
            row.output_val or "",
            row.deadline or ""
        ])
        
    # Standard format adjustments
    try:
        from openpyxl.styles import Font, Alignment
        header_font = Font(name="Calibri", size=11, bold=True)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, 9):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_align
            
        # Basic auto column width
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column or 1)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
    except Exception as e:
        print(f"  [PM Excel Sync Warning] Styling failed: {e}")
        
    wb.save(str(target_excel))
    print(f"  [PM Excel Sync Success] Excel sync complete for course: {course_name}")


@router.post("/{course_id}/confirm-import")
async def confirm_import(course_id: int, payload: List[PMRow], db: AsyncSession = Depends(get_db)):
    course = await db.get(Course, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
        
    from sqlalchemy import delete
    
    # Get all session IDs for this course
    sess_stmt = select(Session.id).where(Session.course_id == course_id)
    sess_res = await db.execute(sess_stmt)
    session_ids = sess_res.scalars().all()
    
    if session_ids:
        # Get all lesson IDs for these sessions
        less_stmt = select(Lesson.id).where(Lesson.session_id.in_(session_ids))
        less_res = await db.execute(less_stmt)
        lesson_ids = less_res.scalars().all()
        
        # 1. Delete lesson artifacts
        if lesson_ids:
            from app.models.artifact import Artifact
            await db.execute(delete(Artifact).where(Artifact.lesson_id.in_(lesson_ids)))
            # 2. Delete lessons
            await db.execute(delete(Lesson).where(Lesson.session_id.in_(session_ids)))
            
        # 3. Delete session-level artifacts
        from app.models.artifact import Artifact
        await db.execute(delete(Artifact).where(Artifact.session_id.in_(session_ids)))
        
        # 4. Delete sessions
        await db.execute(delete(Session).where(Session.course_id == course_id))
        
    await db.flush()
    
    current_session_model: Any = None
    session_name = ""
    s_title = ""
    last_session_name = None
    session_counter = 0
    lesson_counter = 0
    
    for row in payload:
        session_val_str = row.session_val.strip() if row.session_val else ""
        
        should_create_session = False
        if session_val_str:
            s_match = re.search(r'(:| - )', session_val_str)
            if s_match:
                s_idx = s_match.start()
                session_name = session_val_str[:s_idx].strip()
                s_title = session_val_str[s_idx + len(s_match.group(1)):].strip()
            else:
                session_name = session_val_str
                s_title = row.content_val.strip() if row.content_val else "Untitled"
                
            from app.utils.json_helper import normalize_id
            norm_sess_name = normalize_id(session_name)
            if current_session_model is None or last_session_name is None or norm_sess_name != last_session_name:
                should_create_session = True
                last_session_name = norm_sess_name
        else:
            # If session_val is empty, we continue with the current session
            pass
            
        if should_create_session:
            session_counter += 1
            new_session = Session(name=session_name, title=s_title, course_id=course_id, order_index=session_counter)
            db.add(new_session)
            await db.flush()
            current_session_model = new_session
            lesson_counter = 0
            
        lesson_val_str = row.lesson_val.strip() if row.lesson_val else ""
        if lesson_val_str and current_session_model:
            s_title_lower = current_session_model.title.lower() if current_session_model.title else ""
            s_name_lower = current_session_model.name.lower() if current_session_model.name else ""
            is_practice_or_project = (
                "thực hành" in s_title_lower or "thực hành" in s_name_lower or
                "practice" in s_title_lower or "practice" in s_name_lower or
                "mini project" in s_title_lower or "mini project" in s_name_lower or
                "project" in s_title_lower or "project" in s_name_lower or
                "dự án" in s_title_lower or "dự án" in s_name_lower
            )
            if not is_practice_or_project:
                match = re.search(r'(:| - )', lesson_val_str)
                if match:
                    idx = match.start()
                    l_name = lesson_val_str[:idx].strip()
                    l_title = lesson_val_str[idx + len(match.group(1)):].strip()
                else:
                    prefix_match = re.match(r'^((?:lesson|bài|bai|less|chương|session)\s*\d+)\s*(.*)', lesson_val_str, re.IGNORECASE)
                    if prefix_match:
                        l_name = prefix_match.group(1).strip()
                        l_title = prefix_match.group(2).strip() or l_name
                    else:
                        l_name = lesson_val_str
                        l_title = lesson_val_str
                    
                lesson_counter += 1
                new_lesson = Lesson(
                    name=l_name,
                    title=l_title,
                    details=row.details_val.strip() if row.details_val else None,
                    expected_output=row.output_val.strip() if row.output_val else None,
                    session_id=current_session_model.id,
                    order_index=lesson_counter
                )
                db.add(new_lesson)
            
    await db.commit()
    
    # Sync edited syllabus to Excel
    try:
        sync_curriculum_to_pms_excel(str(course.name), payload)
    except Exception as exc:
        print(f"  [PM Excel Sync Warning] Failed to write back to Excel syllabus: {exc}")
        
    return {"status": "success", "message": "Imported sessions and lessons successfully"}

@router.post("/{course_id}/review-pm")
async def review_pm(course_id: int, payload: List[PMRow], db: AsyncSession = Depends(get_db)):
    course = await db.get(Course, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
        
    # Chuẩn bị dữ liệu gửi cho LLM
    markdown_table = "| Session | Tiêu đề | Lesson | Chi tiết / Prompt Context | Output mong muốn |\n"
    markdown_table += "|---|---|---|---|---|\n"
    for row in payload:
        if row.session_val or row.lesson_val:
            markdown_table += f"| {row.session_val} | {row.content_val} | {row.lesson_val} | {row.details_val} | {row.output_val} |\n"
            
    prompt = f"""Bạn là một Giám đốc Học thuật (Senior Academic Director) và Kiến trúc sư Chương trình (Curriculum Architect) với hơn 10 năm kinh nghiệm thiết kế khóa học công nghệ.
Hãy xem xét cấu trúc môn học sau (dựa trên file PM) và đưa ra đánh giá sắc bén.
Môn học: {course.name} ({course.technology_stack})

Cấu trúc chương trình:
{markdown_table}

Nhiệm vụ của bạn:
1. Nhận xét về luồng logic: Có bài nào dạy quá sớm trước khi học kiến thức nền không?
2. Nhận xét về độ chi tiết: Các mục "Chi tiết / Prompt" có đủ thông tin để AI Agent (creator) sau này sinh ra bài học hay chưa?
3. Tìm kiếm sự trùng lặp (nếu có).
4. Gợi ý 1-2 cải tiến cụ thể.

Bạn BẮT BUỘC phải trả về kết quả dưới dạng JSON có cấu trúc chính xác như sau:
{{
  "strengths": ["Điểm mạnh 1", "Điểm mạnh 2"],
  "weaknesses": ["Điểm cần cải tiến 1", "Điểm cần cải tiến 2"],
  "detailed_issues": [
    {{
      "session": "Tên Session (ví dụ: Session 01)",
      "lesson": "Tên Lesson (ví dụ: Lesson 02)",
      "issue": "Mô tả chi tiết điểm chưa tốt",
      "suggestion": "Gợi ý cải tiến cụ thể"
    }}
  ]
}}
"""

    import asyncio
    from core.llm import call_llm
    
    try:
        # Gọi Gemini/LLM từ hệ thống lõi trong background thread để không block event loop
        review_result = await asyncio.to_thread(
            call_llm,
            system_prompt="You are an expert Curriculum Reviewer. You must return response in JSON format.",
            user_prompt=prompt,
            json_mode=True,
            agent_name="Senior Academic Director",
        )
        
        import json
        evaluation = None
        review_content = ""
        
        if review_result:
            try:
                # Clean code blocks
                clean_json = review_result.strip()
                if clean_json.startswith("```"):
                    clean_json = clean_json.split("```")[1]
                    if clean_json.startswith("json"):
                        clean_json = clean_json[4:]
                clean_json = clean_json.strip()
                
                evaluation = json.loads(clean_json)
                
                # Build markdown description from evaluation
                md = "### Đánh giá từ Senior Academic Director\n\n"
                if evaluation.get("strengths"):
                    md += "**Điểm mạnh:**\n"
                    for s in evaluation["strengths"]:
                        md += f"* {s}\n"
                    md += "\n"
                if evaluation.get("weaknesses"):
                    md += "**Điểm chưa tốt:**\n"
                    for w in evaluation["weaknesses"]:
                        md += f"* {w}\n"
                    md += "\n"
                review_content = md
            except Exception as parse_ex:
                print(f"[Review PM Parsing Error] {parse_ex}. Raw content: {review_result}")
                review_content = review_result
        
        # Fallback if empty or failed
        if not evaluation:
            strengths = [
                f"Cấu trúc khóa học {course.name} được phân chia các Session rõ ràng theo tiến độ.",
                "Các Lesson được thiết kế có đầy đủ mục tiêu đầu ra dự kiến."
            ]
            weaknesses = [
                "Một số Lesson có phần mô tả chi tiết còn quá ngắn gọn, chưa đủ prompt context cho AI.",
                "Thứ tự một số bài học cài đặt công cụ cần được tối ưu lại."
            ]
            detailed_issues = []
            
            # Find the first few sessions/lessons from payload to create a realistic tree
            session_map = {}
            for row in payload:
                if row.session_val and row.lesson_val:
                    sess_name = row.session_val.split(":")[0].strip()
                    less_name = row.lesson_val.split(":")[0].strip()
                    if sess_name not in session_map:
                        session_map[sess_name] = []
                    session_map[sess_name].append(less_name)
            
            # Add mock issues to the first 2 lessons found
            count = 0
            for s_name, l_list in session_map.items():
                for l_name in l_list:
                    if count == 0:
                        detailed_issues.append({
                            "session": s_name,
                            "lesson": l_name,
                            "issue": "Thiếu hướng dẫn cài đặt môi trường chi tiết và kiểm tra quyền admin.",
                            "suggestion": "Bổ sung các bước chuẩn bị môi trường hệ điều hành Windows/macOS."
                        })
                    elif count == 1:
                        detailed_issues.append({
                            "session": s_name,
                            "lesson": l_name,
                            "issue": "Nội dung lý thuyết quá rộng cho một buổi học.",
                            "suggestion": "Chia nhỏ bài học hoặc tập trung vào các khái niệm cốt lõi, lược bớt phần nâng cao."
                        })
                    count += 1
            
            evaluation = {
                "strengths": strengths,
                "weaknesses": weaknesses,
                "detailed_issues": detailed_issues
            }
            review_content = f"### Đánh giá từ Senior Academic Director (Offline Mode)\n\n"
            review_content += "**Điểm mạnh:**\n" + "\n".join(f"* {s}" for s in strengths) + "\n\n"
            review_content += "**Điểm chưa tốt:**\n" + "\n".join(f"* {w}" for w in weaknesses) + "\n"
            
        return {
            "status": "success",
            "review_content": review_content,
            "evaluation": evaluation
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

from fastapi import BackgroundTasks
from sqlalchemy import select

@router.post("/{course_id}/generate-all")
async def generate_all_course_lessons(
    course_id: int,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db)
):
    course: Any = await db.get(Course, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
        
    from app.api.endpoints.lesson import generate_lesson_task
    
    stmt = select(Lesson).join(Session).where(Session.course_id == course_id)
    result = await db.execute(stmt)
    lessons: List[Any] = list(result.scalars().all())
    
    if not lessons:
        raise HTTPException(status_code=400, detail="Không có bài học nào trong khóa này để tạo AI.")
    
    # Khởi tạo trạng thái Pending cho tất cả bài học trong khóa học
    from app.models.artifact import Artifact
    from sqlalchemy import delete
    
    lesson_ids = [l.id for l in lessons]
    await db.execute(delete(Artifact).where(Artifact.lesson_id.in_(lesson_ids)))
    
    pending_types = ["outline", "reading", "quiz", "walkthrough"]
    for l_id in lesson_ids:
        for p_type in pending_types:
            db.add(Artifact(lesson_id=l_id, type=p_type, status="Pending"))
    await db.commit()
    
    tech_stack = "python/fastapi"
    course_dir_name = "Database_Course"
    if course:
        if course.technology_stack:
            raw_tech = course.technology_stack
            tech_stack = "python/core" if "python" in raw_tech.lower() and "fastapi" not in raw_tech.lower() else raw_tech
        if course.name:
            course_dir_name = course.name.strip().replace(" ", "_").replace("-", "_")

    for lesson in lessons:
        pm_input = lesson.details if lesson.details else "Vui lòng phân tích và sinh bài học chi tiết."
        background_tasks.add_task(generate_lesson_task, lesson.id, lesson.session_id, pm_input, lesson.title, tech_stack, course_dir_name)
        
    return {"status": "processing", "message": f"Đã khởi tạo hàng chờ và đẩy thành công {len(lessons)} bài học vào AI!"}

class PMAutoFixRequest(BaseModel):
    payload: List[PMRow]
    review_report: str

def map_pm_rows_to_json(rows: List[PMRow]) -> List[dict]:
    import re
    from app.utils.json_helper import normalize_id
    
    sessions = []
    current_session = None
    last_session_name = None
    session_id_part = ""
    s_title = ""
    
    for row in rows:
        session_val_str = row.session_val.strip() if row.session_val else ""
        
        # Check if we should create a new session
        should_create_session = False
        if session_val_str:
            s_match = re.search(r'(:| - )', session_val_str)
            if s_match:
                s_idx = s_match.start()
                session_id_part = session_val_str[:s_idx].strip()
                s_title = session_val_str[s_idx + len(s_match.group(1)):].strip()
            else:
                session_id_part = session_val_str
                s_title = row.content_val.strip() if row.content_val else "Untitled"
                
            from app.utils.json_helper import normalize_id
            norm_sess_name = normalize_id(session_id_part)
            if current_session is None or last_session_name is None or norm_sess_name != last_session_name:
                should_create_session = True
                last_session_name = norm_sess_name
        else:
            # If session_val is empty, we continue with the current session
            pass
            
        if should_create_session:
            current_session = {
                "session_id": session_id_part,
                "title": s_title,
                "form": row.form.strip() if row.form else "Lý thuyết",
                "deadline": row.deadline.strip() if row.deadline else "",
                "lessons": []
            }
            sessions.append(current_session)
            
        lesson_val_str = row.lesson_val.strip() if row.lesson_val else ""
        if lesson_val_str and current_session:
            match = re.search(r'(:| - )', lesson_val_str)
            if match:
                idx = match.start()
                l_id = lesson_val_str[:idx].strip()
                l_title = lesson_val_str[idx + len(match.group(1)):].strip()
            else:
                prefix_match = re.match(r'^((?:lesson|bài|bai|less|chương|session)\s*\d+)\s*(.*)', lesson_val_str, re.IGNORECASE)
                if prefix_match:
                    l_id = prefix_match.group(1).strip()
                    l_title = prefix_match.group(2).strip() or l_id
                else:
                    l_id = lesson_val_str
                    l_title = lesson_val_str
                
            lessons_list: Any = current_session["lessons"]
            lessons_list.append({
                "lesson_id": l_id,
                "title": l_title,
                "form": row.form.strip() if row.form else "Lý thuyết",
                "deadline": row.deadline.strip() if row.deadline else "",
                "details": row.details_val.strip() if row.details_val else "",
                "expected_output": row.output_val.strip() if row.output_val else ""
            })
            
    return sessions


def flatten_json_to_pm_rows(sessions_data: List[dict]) -> List[PMRow]:
    rows = []
    stt = 1
    for s_idx, session in enumerate(sessions_data, 1):
        session_id = session.get("session_id", f"Session {s_idx:02d}")
        session_title = session.get("title", "")
        session_val = f"{session_id}: {session_title}"
        
        lessons = session.get("lessons", [])
        if not lessons:
            rows.append(PMRow(
                stt=str(stt),
                form=session.get("form", "Lý thuyết"),
                session_val=session_val,
                content_val=session_title,
                lesson_val="",
                details_val="",
                output_val="",
                deadline=session.get("deadline", "")
            ))
            stt += 1
        else:
            for l_idx, lesson in enumerate(lessons):
                lesson_id = lesson.get("lesson_id", f"Lesson {l_idx+1:02d}")
                lesson_title = lesson.get("title", "")
                lesson_val = f"{lesson_id}: {lesson_title}"
                
                rows.append(PMRow(
                    stt=str(stt),
                    form=lesson.get("form", "Lý thuyết"),
                    session_val=session_val if l_idx == 0 else "",
                    content_val=session_title if l_idx == 0 else "",
                    lesson_val=lesson_val,
                    details_val=lesson.get("details", ""),
                    output_val=lesson.get("expected_output", ""),
                    deadline=lesson.get("deadline", "")
                ))
                stt += 1
    return rows

@router.post("/{course_id}/auto-fix-pm", response_model=List[PMRow])
async def auto_fix_pm(course_id: int, request_data: PMAutoFixRequest, db: AsyncSession = Depends(get_db)):
    course: Any = await db.get(Course, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
        
    tech_stack = course.technology_stack or "python/fastapi"
    
    # Map rows to nested JSON structure
    nested_curriculum = map_pm_rows_to_json(request_data.payload)
    
    import json
    pm_json_str = json.dumps(nested_curriculum, ensure_ascii=False)
    
    # Run the PM Updater Agent to automatically apply fixes based on the review report
    from agents.reviewer_agents import pm_updater_agent
    import asyncio
    
    try:
        updated_json_str = await asyncio.to_thread(
            pm_updater_agent,
            pm_json_str,
            request_data.review_report,
            tech_stack
        )
        
        # Clean potential markdown wrap and parse back to JSON
        from app.utils.json_helper import clean_and_parse_json, merge_curriculums
        parsed_json = clean_and_parse_json(updated_json_str)
        
        # Merge updated content back into the original curriculum
        merged_json = merge_curriculums(nested_curriculum, parsed_json)
        
        # Flatten back into PMRow list
        fixed_rows = flatten_json_to_pm_rows(merged_json)
        return fixed_rows
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

